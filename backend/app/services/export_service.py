import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportService:
    """Export service for generating Excel files"""

    def export_recommendation(self, result: dict, input_snapshot: dict) -> bytes:
        """Generate Excel file for recommendation result

        Args:
            result: {"rush": [...], "stable": [...], "safe": [...]}
            input_snapshot: {"rank": ..., "province": ..., "subject_type": ..., "exam_type": ...}

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "志愿推荐方案"

        # Styles
        title_font = Font(name="Microsoft YaHei", size=16, bold=True)
        subtitle_font = Font(name="Microsoft YaHei", size=10, color="666666")
        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell_font = Font(name="Microsoft YaHei", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Category colors
        category_styles = {
            "冲": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "稳": PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid"),
            "保": PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid"),
        }

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "智愿 - 高考志愿推荐方案"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        ws.row_dimensions[1].height = 30

        # Subtitle with input params
        rank = input_snapshot.get("rank", "")
        province = input_snapshot.get("province", "")
        subject_type = input_snapshot.get("subject_type", "")
        exam_type = input_snapshot.get("exam_type", "普通类")
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M")

        ws.merge_cells("A2:G2")
        ws["A2"] = f"位次: {rank}  |  省份: {province}  |  科类: {subject_type}  |  考试类型: {exam_type}  |  生成时间: {created_at}"
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = center_align
        ws.row_dimensions[2].height = 20

        # Empty row
        ws.row_dimensions[3].height = 10

        # Header row
        headers = ["序号", "类别", "院校名称", "专业名称", "历年最低位次", "位次比", "适配度"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[4].height = 25

        # Data rows
        row_idx = 5
        seq = 1

        for category, items in [("冲", result.get("rush", [])),
                                 ("稳", result.get("stable", [])),
                                 ("保", result.get("safe", []))]:
            for item in items:
                # 序号
                cell = ws.cell(row=row_idx, column=1, value=seq)
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border

                # 类别
                cell = ws.cell(row=row_idx, column=2, value=category)
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = category_styles[category]

                # 院校名称
                cell = ws.cell(row=row_idx, column=3, value=item.get("university_name", ""))
                cell.font = cell_font
                cell.alignment = left_align
                cell.border = thin_border

                # 专业名称
                cell = ws.cell(row=row_idx, column=4, value=item.get("major_name", ""))
                cell.font = cell_font
                cell.alignment = left_align
                cell.border = thin_border

                # 历年最低位次
                min_rank = item.get("min_rank")
                cell = ws.cell(row=row_idx, column=5, value=min_rank if min_rank else "")
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border

                # 位次比
                rank_ratio = item.get("rank_ratio")
                cell = ws.cell(row=row_idx, column=6, value=f"{rank_ratio:.3f}" if rank_ratio else "")
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border

                # 适配度
                adapter_score = item.get("adapter_score")
                cell = ws.cell(row=row_idx, column=7, value=f"{adapter_score:.1f}" if adapter_score else "")
                cell.font = cell_font
                cell.alignment = center_align
                cell.border = thin_border

                ws.row_dimensions[row_idx].height = 22
                row_idx += 1
                seq += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 8   # 序号
        ws.column_dimensions["B"].width = 10  # 类别
        ws.column_dimensions["C"].width = 25  # 院校名称
        ws.column_dimensions["D"].width = 25  # 专业名称
        ws.column_dimensions["E"].width = 15  # 历年最低位次
        ws.column_dimensions["F"].width = 12  # 位次比
        ws.column_dimensions["G"].width = 12  # 适配度

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
